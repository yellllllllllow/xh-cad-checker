# -*- coding: utf-8 -*-
"""
xh-cad-checker · serial_matrix.py
生成「序号矩阵表」：把每个层级从最小号到最大号列出，标出有/缺。

两大部分（同一份 HTML / 同一份 XLSX 的工作簿）：
  1) 概览热力图：每个序号一个色块（绿=有 / 红=缺），密集排布，一眼定位缺号。
  2) 明细矩阵：每格 [序号 | 完整文件名(带扩展名) | 状态]，绿有红缺，按层级分段。

布局：概览默认每行 25 个色块；明细默认每行 3 单元 = 9 列（--units 可调）。
输出：HTML 始终生成（标准库，零依赖）；XLSX 在检测到 openpyxl 时一并生成，
      否则提示 `pip install openpyxl` 并仅出 HTML。全程只读。

用法：
  python serial_matrix.py --dir "E:/智能药仓 4.2B/02-01 固定货架"
  python serial_matrix.py --dir <目录> --out matrix.html --xlsx matrix.xlsx
  python serial_matrix.py --dir <目录> --recursive --units 4 --ov-cols 30
"""
import os
import re
import argparse

try:
    import html as _html
    _esc = _html.escape
except Exception:
    _esc = lambda s: s


def scan(root, recursive):
    items = []
    if recursive:
        for dp, _, fns in os.walk(root):
            for fn in fns:
                if not fn.startswith("~$"):
                    items.append(os.path.join(dp, fn))
    else:
        for fn in os.listdir(root):
            full = os.path.join(root, fn)
            if os.path.isfile(full) and not fn.startswith("~$"):
                items.append(full)
    return items


def collect(root, recursive):
    """level -> seq -> [完整文件名(带扩展名)]"""
    occ = {}
    for full in scan(root, recursive):
        fn = os.path.basename(full)
        if "." not in fn:
            continue
        base, ext = fn.rsplit(".", 1)
        ext = ext.lower()
        if ext in ("sldprt", "sldasm", "slddrw"):
            m = re.match(r"^XH\d+\.(\d+\.\d+)-(\d+)", base)
            if m:
                occ.setdefault(m.group(1), {}).setdefault(int(m.group(2)), []).append(fn)
    return occ


# ----------------------------- HTML -----------------------------
def build_html(occ, units, ov_cols):
    SUB = 3
    COLS = units * SUB
    h = ['<html><head><meta charset="utf-8"><title>序号矩阵</title>']
    ov_rule = '.ov{display:grid;grid-template-columns:repeat(%d,minmax(34px,1fr));gap:3px;margin:6px 0 16px;}' % ov_cols
    style = ('<style>'
             'body{font-family:-apple-system,Segoe UI,Microsoft YaHei,sans-serif;color:#222;background:#fafafa;margin:20px;}'
             'h2{color:#1a3c6e;}'
             'h3{color:#1a3c6e;margin:18px 0 4px;}'
             '.sum{background:#eef4ff;padding:8px 12px;border-left:4px solid #1a3c6e;margin:8px 0 14px;font-size:13px;line-height:1.7;}'
             + ov_rule +
             '.ovcell{padding:4px 2px;text-align:center;border-radius:4px;font-family:Consolas,monospace;font-size:11px;}'
             '.ov-have{background:#dcfce7;color:#166534;}'
             '.ov-miss{background:#fee2e2;color:#991b1b;font-weight:700;}'
             'table{border-collapse:collapse;background:#fff;font-size:11.5px;margin-bottom:8px;}'
             'th,td{border:1px solid #d4d9e2;padding:3px 6px;text-align:left;vertical-align:top;}'
             'th{background:#1a3c6e;color:#fff;}'
             'td.seq{font-family:Consolas,monospace;font-weight:600;text-align:center;}'
             'td.name{font-family:Consolas,monospace;}'
             'td.have{background:#eafaef;color:#1a7a3c;font-weight:700;text-align:center;}'
             'td.miss{background:#fdecec;color:#b00020;font-weight:700;text-align:center;}'
             '.legend{font-size:12px;color:#555;margin:2px 0 10px;}'
             '</style></head><body>')
    h.append(style)
    h.append('<h2>序号矩阵（概览热力图 + 明细）</h2>')
    h.append('<div class="sum">绿块=有文件，红块=缺(无文件)。'
             '上半「概览」密集色块用于一眼定位缺号；下半「明细」每格=序号|完整文件名|状态，带文件名便于改图。</div>')
    for level in sorted(occ.keys()):
        seqs = sorted(occ[level].keys())
        lo, hi = min(seqs), max(seqs)
        nums = list(range(lo, hi + 1))
        have = len(seqs)
        miss = (hi - lo + 1) - have
        h.append('<h3>层级 %s &nbsp; 范围 %03d~%03d &nbsp; 有 %d / 缺 %d</h3>' % (level, lo, hi, have, miss))
        # 概览热力图
        h.append('<div class="legend">▼ 概览（每行 %d 个，红=缺）</div>' % ov_cols)
        h.append('<div class="ov">')
        for n in nums:
            cls = 'ov-have' if n in occ[level] else 'ov-miss'
            h.append('<div class="ovcell %s">%03d</div>' % (cls, n))
        h.append('</div>')
        # 明细矩阵
        h.append('<div class="legend">▼ 明细（每行 %d 单元 = %d 列：序号|完整文件名|状态）</div>' % (units, COLS))
        h.append('<table><tr>')
        for _ in range(units):
            h.append('<th>序号</th><th>完整文件名</th><th>状态</th>')
        h.append('</tr>')
        for i, n in enumerate(nums):
            if i % units == 0:
                h.append('<tr>')
            h.append('<td class="seq">%03d</td>' % n)
            if n in occ[level]:
                names = '<br>'.join(_esc(x) for x in sorted(occ[level][n]))
                h.append('<td class="name">%s</td>' % names)
                h.append('<td class="have">有</td>')
            else:
                h.append('<td class="name"></td><td class="miss">缺</td>')
            if i % units == units - 1:
                h.append('</tr>')
        if len(nums) % units != 0:
            for _ in range(units - (len(nums) % units)):
                h.append('<td></td><td></td><td></td>')
            h.append('</tr>')
        h.append('</table>')
    h.append('</body></html>')
    return '\n'.join(h)


# ----------------------------- XLSX -----------------------------
def build_xlsx(occ, units, ov_cols, out_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except Exception:
        return False, "未安装 openpyxl（pip install openpyxl 后可生成 xlsx）"
    SUB = 3
    COLS = units * SUB
    thin = Side(style='thin', color='D4D9E2')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    have_fill = PatternFill('solid', fgColor='EAF7EF')
    miss_fill = PatternFill('solid', fgColor='FDECEC')
    ov_have = PatternFill('solid', fgColor='DCFCE7')
    ov_miss = PatternFill('solid', fgColor='FEE2E2')
    hdr_fill = PatternFill('solid', fgColor='1A3C6E')
    hdr_font = Font(color='FFFFFF', bold=True)
    have_font = Font(color='1A7A3C', bold=True)
    miss_font = Font(color='B00020', bold=True)
    ov_miss_font = Font(color='991B1B', bold=True)
    seq_font = Font(name='Consolas', bold=True)
    center = Alignment(horizontal='center', vertical='center')
    leftwrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
    wb = Workbook()
    for level in sorted(occ.keys()):
        seqs = sorted(occ[level].keys())
        lo, hi = min(seqs), max(seqs)
        nums = list(range(lo, hi + 1))
        have = len(seqs)
        # --- 概览 sheet ---
        ws = wb.create_sheet(title='层级 %s 概览' % level)
        ws.append(['层级 %s  范围 %03d~%03d  有 %d / 缺 %d（绿=有 红=缺）'
                   % (level, lo, hi, have, (hi - lo + 1) - have)])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ov_cols)
        for i, n in enumerate(nums):
            r = 2 + i // ov_cols
            c = 1 + i % ov_cols
            cell = ws.cell(row=r, column=c, value='%03d' % n)
            cell.alignment = center
            if n in occ[level]:
                cell.fill = ov_have
            else:
                cell.fill = ov_miss
                cell.font = ov_miss_font
        for c in range(1, ov_cols + 1):
            ws.column_dimensions[chr(65 + c - 1)].width = 6
        ws.freeze_panes = 'A2'
        # --- 明细 sheet ---
        wd = wb.create_sheet(title='层级 ' + level)
        wd.append(['层级 %s  范围 %03d~%03d  有 %d / 缺 %d'
                   % (level, lo, hi, have, (hi - lo + 1) - have)])
        wd.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COLS)
        hdr = []
        for _ in range(units):
            hdr += ['序号', '完整文件名', '状态']
        wd.append(hdr)
        for c in range(1, COLS + 1):
            cell = wd.cell(row=2, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center
            cell.border = border
        for i, n in enumerate(nums):
            r = 3 + i // units
            u = i % units
            cseq, cname, cst = 1 + u * SUB, 2 + u * SUB, 3 + u * SUB
            a = wd.cell(row=r, column=cseq, value='%03d' % n)
            a.font = seq_font
            a.alignment = center
            a.border = border
            if n in occ[level]:
                b = wd.cell(row=r, column=cname, value='\n'.join(sorted(occ[level][n])))
                b.alignment = leftwrap
                b.border = border
                st = wd.cell(row=r, column=cst, value='有')
                st.fill = have_fill
                st.font = have_font
            else:
                wd.cell(row=r, column=cname, value='').border = border
                st = wd.cell(row=r, column=cst, value='缺')
                st.fill = miss_fill
                st.font = miss_font
            st.alignment = center
            st.border = border
        for u in range(units):
            wd.column_dimensions[chr(65 + u * SUB)].width = 7
            wd.column_dimensions[chr(65 + u * SUB + 1)].width = 30
            wd.column_dimensions[chr(65 + u * SUB + 2)].width = 6
        wd.freeze_panes = 'A3'
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    wb.save(out_path)
    return True, out_path


def main():
    ap = argparse.ArgumentParser(description="生成序号矩阵表（概览热力图 + 明细）")
    ap.add_argument("--dir", required=True, help="目标目录")
    ap.add_argument("--out", default=os.path.join(os.getcwd(), "serial_matrix.html"),
                    help="HTML 输出路径（默认当前目录 serial_matrix.html）")
    ap.add_argument("--xlsx", default=None, help="XLSX 输出路径（可选；需 openpyxl）")
    ap.add_argument("--units", type=int, default=3, help="明细每行单元数（默认3=9列）")
    ap.add_argument("--ov-cols", type=int, default=25, help="概览每行色块数（默认25）")
    ap.add_argument("--recursive", action="store_true", help="递归子目录")
    args = ap.parse_args()

    if args.units < 1:
        args.units = 1
    if args.ov_cols < 1:
        args.ov_cols = 1
    occ = collect(args.dir, args.recursive)
    if not occ:
        print("[提示] 未在该目录发现带项目代号的 sldprt/sldasm/slddrw 文件。")
    html_txt = build_html(occ, args.units, args.ov_cols)
    with open(args.out, "w", encoding="utf-8") as fh:
        fh.write(html_txt)
    print(f"HTML -> {args.out}  (层级: {', '.join(sorted(occ.keys())) or '无'})")

    if args.xlsx:
        ok, msg = build_xlsx(occ, args.units, args.ov_cols, args.xlsx)
        print(("XLSX -> " + msg) if ok else ("XLSX 跳过: " + msg))


if __name__ == "__main__":
    main()
